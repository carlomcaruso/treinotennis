#!/usr/bin/env python3
"""
Lê a planilha 'Periodização Carlo - Tênis - 2026.xlsx' no Dropbox e regenera data.js.

Uso local (teste, com o arquivo já baixado):
    python3 scripts/sync_dropbox.py --arquivo ~/Downloads/planilha.xlsx

Uso no GitHub Actions (baixa do Dropbox):
    python3 scripts/sync_dropbox.py
Requer as variáveis DROPBOX_APP_KEY, DROPBOX_APP_SECRET, DROPBOX_REFRESH_TOKEN.
"""
import os, sys, json, re, argparse, datetime, unicodedata

CAMINHO_DROPBOX = "/PESSOAL/TREINOS/2025/Periodização Carlo - Tênis - 2026.xlsx"
DIAS   = ['SÁB','DOM','SEG','TER','QUA','QUI','SEX']
TURNOS = ['MANHÃ','ALMOÇO','NOITE']


# ---------------------------------------------------------------- utilidades
def limpa(v):
    if v is None:
        return ''
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()

def chave(v):
    """Normaliza texto para comparar cabeçalhos: sem acento, maiúsculo, sem espaço extra."""
    s = limpa(v).upper()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s).strip()

def para_iso(v):
    """Converte data em ISO YYYY-MM-DD. Aceita datetime, 'M/D/YY' e 'YYYY-MM-DD'."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    s = limpa(v)
    if not s:
        return ''
    if re.fullmatch(r'\d{4}-\d{2}-\d{2}', s):
        return s
    m = re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', s)
    if m:
        mes, dia, ano = (int(x) for x in m.groups())
        if ano < 100:
            ano += 2000
        try:
            return datetime.date(ano, mes, dia).isoformat()
        except ValueError:
            return s          # data inválida na planilha: preserva o original
    return s

def acha_cabecalho(ws, obrigatorios, limite=400):
    """Procura a linha de cabeçalho e devolve (linha, {NOME: indice_coluna})."""
    alvo = {chave(o) for o in obrigatorios}
    for linha in ws.iter_rows(min_row=1, max_row=min(limite, ws.max_row)):
        mapa = {}
        for cel in linha:
            k = chave(cel.value)
            if k:
                mapa.setdefault(k, cel.column)
        if alvo.issubset(mapa.keys()):
            return linha[0].row, mapa
    return None, None


# ---------------------------------------------------------------- abas
def le_partidas(wb):
    ws = next((wb[n] for n in wb.sheetnames if chave(n).startswith('ACOMPANHAMENTO')), None)
    if ws is None:
        raise SystemExit("aba ACOMPANHAMENTO não encontrada")
    linha, cols = acha_cabecalho(ws, ['DATA', 'OPONENTE', 'RESULTADO', 'QUADRA'])
    if not cols:
        raise SystemExit("cabeçalho da aba ACOMPANHAMENTO não encontrado")

    def val(r, nome):
        c = cols.get(chave(nome))
        return ws.cell(row=r, column=c).value if c else None

    partidas, vazias = [], 0
    for r in range(linha + 1, ws.max_row + 1):
        data, op = val(r, 'DATA'), limpa(val(r, 'OPONENTE'))
        if not limpa(data) and not op:
            vazias += 1
            if vazias >= 15:
                break
            continue
        vazias = 0
        if not op:
            continue
        partidas.append({
            'data':     para_iso(data),
            'oponente': op.upper(),
            'res':      limpa(val(r, 'RESULTADO')).upper()[:1],
            'quadra':   limpa(val(r, 'QUADRA')).upper(),
            'placar':   limpa(val(r, 'PLACAR')),
            'qual':     limpa(val(r, 'QUALIDADE')).upper(),
            'conq':     limpa(val(r, 'CONQUISTAS')),
        })
    return partidas


def le_derrotas(wb):
    ws = next((wb[n] for n in wb.sheetnames if chave(n).startswith('DERROTAS')), None)
    if ws is None:
        return []
    linha, cols = acha_cabecalho(ws, ['DATA', 'OPONENTE', 'QUADRA', 'NIVEL'])
    if not cols:
        return []

    def val(r, nome):
        c = cols.get(chave(nome))
        return ws.cell(row=r, column=c).value if c else None

    saida, vazias = [], 0
    for r in range(linha + 1, ws.max_row + 1):
        op = limpa(val(r, 'OPONENTE'))
        if not op:
            vazias += 1
            if vazias >= 10:
                break
            continue
        vazias = 0
        saida.append({
            'data':     para_iso(val(r, 'DATA')),
            'oponente': op.upper(),
            'quadra':   limpa(val(r, 'QUADRA')).upper(),
            'nivel':    limpa(val(r, 'NIVEL')).upper(),
            'vitoria':  para_iso(val(r, 'DATA VITORIA')),
        })
    return saida


def le_semana(wb):
    """Lê o bloco de semana mais recente (o primeiro de cima para baixo)."""
    ws = next((wb[n] for n in wb.sheetnames if chave(n) == 'SEMANA'), None)
    if ws is None:
        return None
    inicio = num = None
    for linha in ws.iter_rows(min_row=1, max_row=min(200, ws.max_row)):
        for cel in linha:
            if chave(cel.value) == 'SEMANA':
                viz = ws.cell(row=cel.row, column=cel.column + 1).value
                if isinstance(viz, (int, float)):
                    inicio, num = cel.row, int(viz)
                    break
        if inicio:
            break
    if not inicio:
        return None

    fase = objetivo = ''
    for r in range(inicio, min(inicio + 6, ws.max_row + 1)):
        for c in range(1, min(6, ws.max_column + 1)):
            k = chave(ws.cell(row=r, column=c).value)
            if k == 'FASE':
                fase = limpa(ws.cell(row=r, column=c + 1).value).upper()
            elif k == 'OBJETIVO':
                objetivo = limpa(ws.cell(row=r, column=c + 1).value).upper()

    grid = {t: [''] * 7 for t in TURNOS}
    objetivos, rotulo = {}, None
    for r in range(inicio, min(inicio + 40, ws.max_row + 1)):
        for c in range(1, min(4, ws.max_column + 1)):
            k = chave(ws.cell(row=r, column=c).value)
            if k in (chave(t) for t in TURNOS):
                turno = next(t for t in TURNOS if chave(t) == k)
                grid[turno] = [limpa(ws.cell(row=r, column=c + 1 + i).value) for i in range(7)]
                # coluna OBJETIVOS fica logo depois dos 7 dias
                rot = limpa(ws.cell(row=r, column=c + 8).value)
                if rot:
                    rotulo = rot.upper()
                    objetivos.setdefault(rotulo, [])
                break
        if rotulo:
            extra = limpa(ws.cell(row=r, column=1).value) or limpa(ws.cell(row=r, column=10).value)
            if extra.startswith('.'):
                objetivos[rotulo].append(extra.lstrip('. ').strip())

    if not any(any(v) for v in grid.values()):
        return None
    return {'numero': num, 'fase': fase or 'PREPARAÇÃO', 'objetivo': objetivo,
            'grid': grid, 'objetivos': objetivos or {}}


# ---------------------------------------------------------------- dropbox
def baixa_do_dropbox(destino):
    import requests
    app_key    = os.environ['DROPBOX_APP_KEY']
    app_secret = os.environ['DROPBOX_APP_SECRET']
    refresh    = os.environ['DROPBOX_REFRESH_TOKEN']

    tk = requests.post('https://api.dropboxapi.com/oauth2/token',
                       data={'grant_type': 'refresh_token', 'refresh_token': refresh},
                       auth=(app_key, app_secret), timeout=30)
    tk.raise_for_status()
    token = tk.json()['access_token']

    caminho = unicodedata.normalize('NFC', CAMINHO_DROPBOX)
    r = requests.post('https://content.dropboxapi.com/2/files/download',
                      headers={'Authorization': f'Bearer {token}',
                               'Dropbox-API-Arg': json.dumps({'path': caminho},
                                                             ensure_ascii=True)},
                      timeout=120)
    r.raise_for_status()
    with open(destino, 'wb') as f:
        f.write(r.content)
    return destino


# ---------------------------------------------------------------- saída
def escreve_data_js(partidas, derrotas, semana, caminho='data.js'):
    anterior = ''
    if os.path.exists(caminho):
        anterior = open(caminho, encoding='utf-8').read()

    # preserva SEMANA_BASE/FASES/BPM do arquivo atual se a leitura da aba falhar
    def bloco(nome, padrao):
        m = re.search(rf'const {nome} = (.*?);\n\n', anterior, re.S)
        return m.group(1) if m else padrao

    semana_js = json.dumps(semana, ensure_ascii=False, indent=0) if semana else bloco(
        'SEMANA_BASE', '{}')

    hoje = datetime.date.today().isoformat()
    txt  = f"// Gerado automaticamente de 'Periodização Carlo - Tênis - 2026.xlsx' em {hoje}\n"
    txt += "// Não edite à mão: o GitHub Actions sobrescreve este arquivo.\n"
    txt += "const HIST = "        + json.dumps(partidas, ensure_ascii=False, indent=0) + ";\n\n"
    txt += "const DERROTAS25 = "  + json.dumps(derrotas, ensure_ascii=False, indent=0) + ";\n\n"
    txt += "const SEMANA_BASE = " + semana_js + ";\n\n"
    txt += "const FASES = "       + bloco('FASES', '{}') + ";\n\n"
    txt += "const BPM = "         + bloco('BPM', '{}') + ";\n\n"
    txt += "const DIAS = "        + json.dumps(DIAS, ensure_ascii=False) + ";\n"
    txt += "const TURNOS = "      + json.dumps(TURNOS, ensure_ascii=False) + ";\n"

    mudou = (txt != anterior)
    if mudou:
        open(caminho, 'w', encoding='utf-8').write(txt)
    return mudou


def carimba_versao(caminho='sw.js'):
    if not os.path.exists(caminho):
        return
    s = open(caminho, encoding='utf-8').read()
    nova = datetime.datetime.now().strftime('%Y-%m-%d-%H%M')
    s = re.sub(r"const VERSION = '.*?';", f"const VERSION = '{nova}';", s, count=1)
    open(caminho, 'w', encoding='utf-8').write(s)
    print(f"sw.js carimbado: {nova}")


# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--arquivo', help='xlsx local (pula o download do Dropbox)')
    ap.add_argument('--dry-run', action='store_true', help='só mostra o resumo, não escreve')
    args = ap.parse_args()

    import openpyxl
    caminho = args.arquivo or baixa_do_dropbox('/tmp/planilha.xlsx')
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=False)

    partidas = le_partidas(wb)
    derrotas = le_derrotas(wb)
    semana   = le_semana(wb)

    v = sum(1 for p in partidas if p['res'] == 'V')
    print(f"partidas: {len(partidas)}  ({v}V / {len(partidas)-v}D)")
    print(f"derrotas 2025: {len(derrotas)}")
    print(f"semana: {semana['numero'] if semana else 'não lida (mantém a anterior)'}")

    if not partidas:
        raise SystemExit("ERRO: nenhuma partida lida — abortando para não apagar o data.js")

    if args.dry_run:
        print("\n-- últimas 3 --")
        for p in partidas[-3:]:
            print(' ', p)
        return

    if escreve_data_js(partidas, derrotas, semana):
        carimba_versao()
        print("MUDOU=1")
    else:
        print("MUDOU=0  (nada novo na planilha)")


if __name__ == '__main__':
    main()
